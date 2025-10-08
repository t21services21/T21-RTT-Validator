"""
Excel Export Module for T21 RTT Validator
Generates downloadable Excel files with validation results
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def create_validation_excel(validation_result, excel_data):
    """
    Create Excel file with validation results
    Returns: BytesIO object ready for download
    """
    wb = Workbook()
    
    # Create sheets
    ws_summary = wb.active
    ws_summary.title = "Validation Summary"
    ws_checklist = wb.create_sheet("Checklist")
    ws_gaps = wb.create_sheet("Gaps Found")
    ws_pas_updates = wb.create_sheet("PAS Updates")
    
    # ===================================
    # SHEET 1: VALIDATION SUMMARY
    # ===================================
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws_summary['A1'] = 'T21 RTT PATHWAY VALIDATION REPORT'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    # Validation Details
    row = 3
    ws_summary[f'A{row}'] = 'VALIDATION DETAILS'
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 1
    details = [
        ['Validator Name:', excel_data.get('Validator_Name', '')],
        ['Validation Date:', excel_data.get('Validation_Date', '')],
        ['RTT Code Assigned:', validation_result.get('RTT_Code', '')],
        ['Clock Status:', excel_data.get('Clock_Status', '')],
        ['Outcome:', excel_data.get('Outcome', '')],
        ['']
    ]
    
    for detail in details:
        ws_summary[f'A{row}'] = detail[0]
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = detail[1]
        row += 1
    
    # Validation Results
    row += 1
    ws_summary[f'A{row}'] = 'VALIDATION RESULTS'
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    
    val_summary = validation_result.get('Validation_Summary', {})
    row += 1
    results = [
        ['Overall Status:', val_summary.get('Overall_Status', '')],
        ['Compliance Rate:', val_summary.get('Compliance_Rate', '')],
        ['Total Actions Required:', val_summary.get('Total_Actions_Required', 0)],
        ['Actions Completed:', val_summary.get('Actions_Completed', 0)],
        ['Actions Outstanding:', val_summary.get('Actions_Outstanding', 0)],
        ['Excel Flag:', val_summary.get('Excel_Flag_Color', '')]
    ]
    
    for result in results:
        ws_summary[f'A{row}'] = result[0]
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = result[1]
        
        # Color code the flag
        if result[0] == 'Excel Flag:':
            flag_color = result[1]
            if flag_color == 'GREEN':
                ws_summary[f'B{row}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif flag_color == 'AMBER':
                ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            elif flag_color == 'RED':
                ws_summary[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        row += 1
    
    # Validation Comments
    row += 1
    ws_summary[f'A{row}'] = 'VALIDATION COMMENTS'
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    ws_summary[f'A{row}'] = excel_data.get('Validation_Comments', '')
    ws_summary.merge_cells(f'A{row}:D{row}')
    ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # ===================================
    # SHEET 2: VALIDATION CHECKLIST
    # ===================================
    
    # Headers
    ws_checklist['A1'] = 'ACTIONS REQUIRED'
    ws_checklist['A1'].font = header_font
    ws_checklist['A1'].fill = header_fill
    
    ws_checklist['B1'] = 'STATUS'
    ws_checklist['B1'].font = header_font
    ws_checklist['B1'].fill = header_fill
    
    compliance = validation_result.get('Action_Compliance', {})
    
    row = 2
    for action in compliance.get('Actions_Required', []):
        ws_checklist[f'A{row}'] = action
        ws_checklist[f'B{row}'] = 'Required'
        ws_checklist[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1
    
    for action in compliance.get('Actions_Reported_In_PAS', []):
        ws_checklist[f'A{row}'] = action
        ws_checklist[f'B{row}'] = 'Completed'
        ws_checklist[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1
    
    # Auto-fit columns
    ws_checklist.column_dimensions['A'].width = 80
    ws_checklist.column_dimensions['B'].width = 15
    
    # ===================================
    # SHEET 3: GAPS FOUND
    # ===================================
    
    ws_gaps['A1'] = 'GAP DESCRIPTION'
    ws_gaps['A1'].font = header_font
    ws_gaps['A1'].fill = header_fill
    
    ws_gaps['B1'] = 'PRIORITY'
    ws_gaps['B1'].font = header_font
    ws_gaps['B1'].fill = header_fill
    
    row = 2
    priority = compliance.get('Priority', 'Medium')
    for gap in compliance.get('Gaps', []):
        ws_gaps[f'A{row}'] = gap
        ws_gaps[f'B{row}'] = priority
        
        # Color code by priority
        if priority == 'High':
            ws_gaps[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif priority == 'Medium':
            ws_gaps[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            ws_gaps[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        row += 1
    
    ws_gaps.column_dimensions['A'].width = 80
    ws_gaps.column_dimensions['B'].width = 15
    
    # ===================================
    # SHEET 4: PAS UPDATES REQUIRED
    # ===================================
    
    ws_pas_updates['A1'] = 'PAS UPDATE ACTION'
    ws_pas_updates['A1'].font = header_font
    ws_pas_updates['A1'].fill = header_fill
    
    row = 2
    for update in validation_result.get('PAS_Update', []):
        ws_pas_updates[f'A{row}'] = update
        row += 1
    
    ws_pas_updates.column_dimensions['A'].width = 100
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def create_batch_results_excel(batch_results):
    """
    Create Excel file with batch validation results
    Returns: BytesIO object ready for download
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Validation Results"
    
    # Headers (matching Excel tracker)
    headers = [
        'Patient Name', 'NHS Number', 'Validator Name', 'Clock Status', 
        'Outcome', 'Validation Date', 'RTT Code', 'Compliance Rate',
        'Flag Color', 'Validation Comments'
    ]
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_num, result in enumerate(batch_results, 2):
        excel_data = result.get('Excel_Report', {})
        val_summary = result.get('Validation_Summary', {})
        
        ws.cell(row=row_num, column=1).value = result.get('patient_name', '')
        ws.cell(row=row_num, column=2).value = result.get('nhs_number', '')
        ws.cell(row=row_num, column=3).value = excel_data.get('Validator_Name', '')
        ws.cell(row=row_num, column=4).value = excel_data.get('Clock_Status', '')
        ws.cell(row=row_num, column=5).value = excel_data.get('Outcome', '')
        ws.cell(row=row_num, column=6).value = excel_data.get('Validation_Date', '')
        ws.cell(row=row_num, column=7).value = result.get('RTT_Code', '')
        ws.cell(row=row_num, column=8).value = val_summary.get('Compliance_Rate', '')
        ws.cell(row=row_num, column=9).value = val_summary.get('Excel_Flag_Color', '')
        ws.cell(row=row_num, column=10).value = excel_data.get('Validation_Comments', '')
        
        # Color code flag column
        flag_cell = ws.cell(row=row_num, column=9)
        flag_color = val_summary.get('Excel_Flag_Color', '')
        if flag_color == 'GREEN':
            flag_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif flag_color == 'AMBER':
            flag_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif flag_color == 'RED':
            flag_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
